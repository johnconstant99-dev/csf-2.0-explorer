#!/usr/bin/env python3
"""
NIST Cybersecurity Framework (CSF) 2.0 Reference & Self-Assessment Tool
Full-stack interactive web app built with Streamlit + SQLite
"""

import streamlit as st
import pandas as pd
import openpyxl
import sqlite3
import re
from datetime import datetime
from pathlib import Path
import base64

# ====================== CONFIG ======================
st.set_page_config(
    page_title="NIST CSF 2.0 | Reference & Assessment",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# DATA_PATH is resolved relative to this script for portability (works in sandbox or when you download the folder)
SCRIPT_DIR = Path(__file__).parent.resolve()
DATA_PATH = SCRIPT_DIR / "cprt_CSF_2_0_0_06-17-2026.xlsx"
DB_PATH = SCRIPT_DIR / "csf_assessments.db"
ASSETS_DIR = SCRIPT_DIR

# Color scheme for Functions (inspired by NIST)
FUNCTION_COLORS = {
    "GV": {"bg": "#5B2C6F", "text": "white", "name": "GOVERN"},
    "ID": {"bg": "#1E8449", "text": "white", "name": "IDENTIFY"},
    "PR": {"bg": "#2874A6", "text": "white", "name": "PROTECT"},
    "DE": {"bg": "#B9770E", "text": "white", "name": "DETECT"},
    "RS": {"bg": "#922B21", "text": "white", "name": "RESPOND"},
    "RC": {"bg": "#1F618D", "text": "white", "name": "RECOVER"},
}

STATUS_LABELS = {
    0: "⚪ Not Assessed",
    1: "🔴 Not Implemented",
    2: "🟡 Partially Implemented",
    3: "🟢 Largely Implemented",
    4: "✅ Fully Implemented",
}

STATUS_COLORS = {
    0: "#6c757d",
    1: "#dc3545",
    2: "#ffc107",
    3: "#28a745",
    4: "#198754",
}

# ====================== DATA LOADING & PARSING ======================
@st.cache_data(show_spinner="Loading NIST CSF 2.0 Core...")
def load_and_parse_csf():
    """Parse the hierarchical CSF 2.0 Excel into a clean DataFrame."""
    wb = openpyxl.load_workbook(DATA_PATH)
    ws = wb["CSF 2.0"]

    records = []
    current_func_id = None
    current_func_name = None
    current_func_desc = None
    current_cat_id = None
    current_cat_name = None
    current_cat_desc = None

    for row_idx in range(3, ws.max_row + 1):
        func_raw = ws.cell(row_idx, 1).value
        cat_raw = ws.cell(row_idx, 2).value
        sub_raw = ws.cell(row_idx, 3).value
        impl = ws.cell(row_idx, 4).value or ""
        refs = ws.cell(row_idx, 5).value or ""

        if func_raw:
            func_raw = str(func_raw).strip()
            func_match = re.match(r"^([A-Z]+) \(([A-Z]{2})\):\s*(.+)$", func_raw)
            if func_match:
                current_func_name = func_match.group(1)
                current_func_id = func_match.group(2)
                current_func_desc = func_match.group(3).strip()
                current_cat_id = None
                current_cat_name = None
                current_cat_desc = None
            continue

        if cat_raw:
            cat_raw = str(cat_raw).strip()
            cat_match = re.match(r"^(.+?) \(([A-Z]{2}\.[A-Z]{2})\):\s*(.+)$", cat_raw)
            if cat_match:
                current_cat_name = cat_match.group(1).strip()
                current_cat_id = cat_match.group(2)
                current_cat_desc = cat_match.group(3).strip()
            continue

        if sub_raw:
            sub_raw = str(sub_raw).strip()
            sub_match = re.match(r"^([A-Z]{2}\.[A-Z]{2}-\d+):\s*(.+)$", sub_raw)
            if sub_match:
                sub_id = sub_match.group(1)
                sub_desc = sub_match.group(2).strip()

                if not current_func_id:
                    current_func_id = sub_id.split(".")[0]
                if not current_cat_id:
                    current_cat_id = sub_id.rsplit("-", 1)[0] if "-" in sub_id else sub_id

                records.append(
                    {
                        "function_id": current_func_id,
                        "function_name": current_func_name or current_func_id,
                        "function_desc": current_func_desc or "",
                        "category_id": current_cat_id,
                        "category_name": current_cat_name or current_cat_id,
                        "category_desc": current_cat_desc or "",
                        "subcategory_id": sub_id,
                        "subcategory_desc": sub_desc,
                        "implementation_examples": impl.strip(),
                        "informative_references": refs.strip(),
                        "is_withdrawn": "Withdrawn" in sub_desc or "withdrawn" in sub_desc.lower(),
                    }
                )

    df = pd.DataFrame(records)

    # Add helper columns
    df["function_label"] = df["function_id"] + " — " + df["function_name"]
    df["short_desc"] = df["subcategory_desc"].str[:120] + "..."

    return df


def get_intro_text():
    wb = openpyxl.load_workbook(DATA_PATH)
    ws = wb["Introduction"]
    intro = {}
    for row in range(1, 6):
        key = ws.cell(row, 1).value
        val = ws.cell(row, 2).value
        if key:
            intro[key] = val
    return intro


# ====================== DATABASE (SQLite) ======================
def init_db():
    """Initialize SQLite DB for assessments and org profile."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Assessments table
    c.execute(
        """
        CREATE TABLE IF NOT EXISTS assessments (
            subcategory_id TEXT PRIMARY KEY,
            status INTEGER DEFAULT 0,
            notes TEXT DEFAULT '',
            updated_at TEXT
        )
    """
    )

    # Org profile / metadata
    c.execute(
        """
        CREATE TABLE IF NOT EXISTS profile (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            org_name TEXT DEFAULT 'My Organization',
            assessment_date TEXT,
            overall_notes TEXT DEFAULT '',
            updated_at TEXT
        )
    """
    )
    # Seed default profile if empty
    c.execute("INSERT OR IGNORE INTO profile (id, org_name, updated_at) VALUES (1, 'My Organization', ?)", 
              (datetime.now().isoformat(),))
    conn.commit()
    conn.close()


def get_assessments_df():
    """Return all assessments as DataFrame (or empty)."""
    conn = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql_query("SELECT * FROM assessments", conn)
    except Exception:
        df = pd.DataFrame(columns=["subcategory_id", "status", "notes", "updated_at"])
    conn.close()
    return df


def get_profile():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM profile WHERE id=1", conn)
    conn.close()
    return df.iloc[0].to_dict() if len(df) > 0 else {"org_name": "My Organization", "overall_notes": "", "assessment_date": None}


def save_assessment(subcategory_id: str, status: int, notes: str):
    """Upsert a single assessment."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    now = datetime.now().isoformat()
    c.execute(
        """
        INSERT INTO assessments (subcategory_id, status, notes, updated_at)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(subcategory_id) DO UPDATE SET
            status=excluded.status,
            notes=excluded.notes,
            updated_at=excluded.updated_at
    """,
        (subcategory_id, status, notes, now),
    )
    conn.commit()
    conn.close()
    st.toast(f"Assessment saved for {subcategory_id}", icon="✅")


def save_profile(org_name: str, overall_notes: str):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    now = datetime.now().isoformat()
    c.execute(
        """
        UPDATE profile SET org_name=?, overall_notes=?, updated_at=? WHERE id=1
    """,
        (org_name, overall_notes, now),
    )
    conn.commit()
    conn.close()


def get_assessment_for_sub(sub_id: str):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT status, notes FROM assessments WHERE subcategory_id=?", (sub_id,))
    row = c.fetchone()
    conn.close()
    if row:
        return {"status": row[0], "notes": row[1]}
    return {"status": 0, "notes": ""}


def get_implementation_stats(csf_df: pd.DataFrame):
    """Merge assessments and compute progress stats."""
    assess_df = get_assessments_df()
    if assess_df.empty:
        merged = csf_df.copy()
        merged["status"] = 0
        merged["notes"] = ""
    else:
        merged = csf_df.merge(
            assess_df[["subcategory_id", "status", "notes"]],
            on="subcategory_id",
            how="left",
        )
        merged["status"] = merged["status"].fillna(0).astype(int)
        merged["notes"] = merged["notes"].fillna("")

    # Calculate progress
    total = len(merged)
    assessed = (merged["status"] > 0).sum()
    fully = (merged["status"] == 4).sum()
    avg_status = merged["status"].mean() if total > 0 else 0
    progress_pct = (avg_status / 4) * 100

    per_function = (
        merged.groupby("function_id")
        .agg(
            subcats=("subcategory_id", "count"),
            assessed=("status", lambda x: (x > 0).sum()),
            avg_status=("status", "mean"),
        )
        .reset_index()
    )
    per_function["progress_pct"] = (per_function["avg_status"] / 4 * 100).round(1)

    return {
        "merged": merged,
        "total": total,
        "assessed": assessed,
        "fully_implemented": fully,
        "progress_pct": round(progress_pct, 1),
        "per_function": per_function,
    }


# ====================== UI HELPERS ======================
def colored_function_badge(func_id: str):
    colors = FUNCTION_COLORS.get(func_id, {"bg": "#555", "text": "white"})
    return f"""
    <span style="
        background-color: {colors['bg']};
        color: {colors['text']};
        padding: 4px 10px;
        border-radius: 6px;
        font-size: 0.85rem;
        font-weight: 600;
        letter-spacing: 0.5px;
    ">{func_id}</span>
    """


def status_badge(status: int):
    label = STATUS_LABELS.get(status, "Unknown")
    color = STATUS_COLORS.get(status, "#6c757d")
    return f"""
    <span style="
        background-color: {color}20;
        color: {color};
        padding: 3px 8px;
        border-radius: 9999px;
        font-size: 0.75rem;
        font-weight: 600;
        border: 1px solid {color}40;
    ">{label}</span>
    """


def render_subcategory_card(row, show_assessment=True):
    """Render a nice card for one subcategory."""
    func_colors = FUNCTION_COLORS.get(row["function_id"], {"bg": "#555", "text": "white"})

    with st.container(border=True):
        cols = st.columns([0.8, 4, 1.2])
        with cols[0]:
            st.markdown(colored_function_badge(row["function_id"]), unsafe_allow_html=True)
            st.caption(row["category_id"])

        with cols[1]:
            st.markdown(f"**{row['subcategory_id']}**")
            st.write(row["subcategory_desc"])

            if row["implementation_examples"]:
                with st.expander("Implementation Examples", expanded=False):
                    st.markdown(row["implementation_examples"].replace("\n", "\n\n"))

            if row.get("informative_references"):
                with st.expander("Informative References", expanded=False):
                    st.code(row["informative_references"], language=None)

        with cols[2]:
            if show_assessment:
                assessment = get_assessment_for_sub(row["subcategory_id"])
                st.markdown(status_badge(assessment["status"]), unsafe_allow_html=True)
                if assessment["notes"]:
                    st.caption(f"📝 {assessment['notes'][:60]}..." if len(assessment["notes"]) > 60 else f"📝 {assessment['notes']}")


# ====================== MAIN APP ======================
def main():
    init_db()
    csf_df = load_and_parse_csf()
    intro = get_intro_text()
    stats = get_implementation_stats(csf_df)
    profile = get_profile()

    # ========== SIDEBAR ==========
    with st.sidebar:
        st.markdown("## 🛡️ NIST CSF 2.0")
        st.caption("Cybersecurity Framework Reference & Assessment Tool")

        st.divider()

        nav = st.radio(
            "Navigate",
            ["📊 Overview", "🔍 Browse & Search", "📝 Self-Assessment", "📈 Reports & Export", "ℹ️ About"],
            label_visibility="collapsed",
        )

        st.divider()

        # Quick profile edit
        with st.expander("🏢 Organization Profile", expanded=False):
            new_org = st.text_input("Organization Name", value=profile["org_name"], key="org_name_input")
            new_notes = st.text_area("Overall Notes / Scope", value=profile.get("overall_notes", ""), height=80, key="org_notes")
            if st.button("Save Profile", use_container_width=True):
                save_profile(new_org, new_notes)
                st.success("Profile updated!")
                st.rerun()

        st.caption(f"Data generated: {intro.get('Generated Date', 'N/A')} | Final version")

    # ========== HEADER ==========
    st.markdown(
        """
        <div style="display:flex; align-items:center; gap:12px; margin-bottom:0.5rem;">
            <h1 style="margin:0; font-size:2.1rem;">NIST Cybersecurity Framework 2.0</h1>
            <span style="background:#0d6efd; color:white; padding:4px 12px; border-radius:8px; font-size:0.9rem; font-weight:600;">v2.0</span>
        </div>
        <p style="color:#555; margin-top:0;">Explore the Core • Assess your posture • Track implementation progress</p>
        """,
        unsafe_allow_html=True,
    )

    # ========== PAGE: OVERVIEW ==========
    if nav == "📊 Overview":
        st.subheader("Executive Dashboard")

        # KPI cards
        kpi1, kpi2, kpi3, kpi4 = st.columns(4)
        kpi1.metric("Functions", "6", help="GOVERN, IDENTIFY, PROTECT, DETECT, RESPOND, RECOVER")
        kpi2.metric("Categories", csf_df["category_id"].nunique())
        kpi3.metric("Subcategories", f"{stats['total']} ({stats['total'] - 79} core + 79 withdrawn)")
        kpi4.metric("Your Progress", f"{stats['progress_pct']}%", f"{stats['assessed']}/{stats['total']} assessed")

        st.divider()

        # Function cards
        st.markdown("### The Six CSF Functions")
        func_cols = st.columns(3)
        func_order = ["GV", "ID", "PR", "DE", "RS", "RC"]

        for idx, fid in enumerate(func_order):
            fdata = csf_df[csf_df["function_id"] == fid].iloc[0]
            fstat = stats["per_function"][stats["per_function"]["function_id"] == fid].iloc[0] if len(stats["per_function"]) > 0 else None

            with func_cols[idx % 3]:
                color = FUNCTION_COLORS[fid]
                with st.container(border=True):
                    st.markdown(
                        f"<div style='background:{color['bg']}; color:{color['text']}; padding:8px 14px; border-radius:8px 8px 0 0;'><strong>{fid} — {color['name']}</strong></div>",
                        unsafe_allow_html=True,
                    )
                    st.caption(fdata["function_desc"][:180] + "…")
                    n_subs = len(csf_df[csf_df["function_id"] == fid])
                    n_cats = csf_df[csf_df["function_id"] == fid]["category_id"].nunique()
                    st.markdown(f"**{n_cats}** Categories • **{n_subs}** Subcategories")

                    if fstat is not None:
                        prog = fstat["progress_pct"]
                        st.progress(prog / 100, text=f"{prog}% implemented (avg)")

                    if st.button(f"Explore {fid}", key=f"explore_{fid}", use_container_width=True):
                        st.session_state["nav_to_browse"] = fid
                        st.switch_page("csf_2_0_explorer.py")  # hack to simulate nav, but since single file use radio change? For simplicity rerun with query
                        # Actually since single page, we'll handle via session or just note

        st.info("💡 Tip: Use the sidebar navigation to dive deeper into Browse, Assessment, or Reports.")

        # Quick withdrawn note
        with st.expander("Note on Withdrawn Subcategories (79 total)"):
            st.write(
                "In CSF 2.0 many subcategories from v1.1 (especially in IDENTIFY) were withdrawn and incorporated into the new GOVERN function or other areas. "
                "They remain in this reference for historical traceability."
            )

    # ========== PAGE: BROWSE & SEARCH ==========
    elif nav == "🔍 Browse & Search":
        st.subheader("Browse the CSF 2.0 Core")

        # Filters
        filter_col1, filter_col2, filter_col3 = st.columns([1.5, 2, 2])
        with filter_col1:
            selected_funcs = st.multiselect(
                "Functions",
                options=csf_df["function_id"].unique().tolist(),
                default=csf_df["function_id"].unique().tolist(),
                format_func=lambda x: f"{x} — {FUNCTION_COLORS.get(x,{}).get('name',x)}",
            )
        with filter_col2:
            search_term = st.text_input("🔎 Search in descriptions, examples, or IDs", placeholder="e.g. supply chain, incident, encryption, GV.OC")
        with filter_col3:
            show_withdrawn = st.checkbox("Include withdrawn subcategories", value=False)

        # Apply filters
        filtered = csf_df[csf_df["function_id"].isin(selected_funcs)]
        if not show_withdrawn:
            filtered = filtered[~filtered["is_withdrawn"]]

        if search_term:
            mask = (
                filtered["subcategory_id"].str.contains(search_term, case=False, na=False)
                | filtered["subcategory_desc"].str.contains(search_term, case=False, na=False)
                | filtered["implementation_examples"].str.contains(search_term, case=False, na=False)
                | filtered["category_name"].str.contains(search_term, case=False, na=False)
            )
            filtered = filtered[mask]

        st.caption(f"Showing **{len(filtered)}** subcategories")

        # Hierarchical or flat view toggle
        view_mode = st.radio("View mode", ["Hierarchical (by Category)", "Flat Table"], horizontal=True, index=0)

        if view_mode == "Hierarchical (by Category)":
            # Group by function then category
            for func_id in selected_funcs:
                func_df = filtered[filtered["function_id"] == func_id]
                if func_df.empty:
                    continue
                color = FUNCTION_COLORS[func_id]
                with st.expander(f"{func_id} — {color['name']} ({len(func_df)} subcategories)", expanded=(len(selected_funcs) == 1)):
                    for cat_id in func_df["category_id"].unique():
                        cat_df = func_df[func_df["category_id"] == cat_id]
                        cat_name = cat_df["category_name"].iloc[0]
                        with st.expander(f"📁 {cat_id} — {cat_name} ({len(cat_df)})", expanded=False):
                            for _, row in cat_df.iterrows():
                                render_subcategory_card(row, show_assessment=True)
                                st.divider()
        else:
            # Flat table
            display_df = filtered[["function_id", "category_id", "subcategory_id", "subcategory_desc", "is_withdrawn"]].copy()
            display_df["status"] = display_df["subcategory_id"].apply(lambda x: get_assessment_for_sub(x)["status"])
            display_df["Status"] = display_df["status"].map(STATUS_LABELS)
            display_df = display_df.drop(columns=["status", "is_withdrawn"])

            st.dataframe(
                display_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "function_id": st.column_config.TextColumn("Function", width="small"),
                    "category_id": st.column_config.TextColumn("Category", width="small"),
                    "subcategory_id": st.column_config.TextColumn("Subcategory", width="medium"),
                    "subcategory_desc": st.column_config.TextColumn("Description", width="large"),
                    "Status": st.column_config.TextColumn("Your Status", width="medium"),
                },
            )

        # Detail + Quick Assess panel
        st.divider()
        st.markdown("### 🎯 Quick Assessment Panel")

        all_ids = filtered["subcategory_id"].tolist()
        if all_ids:
            selected_sub = st.selectbox(
                "Select a Subcategory to view details & update assessment",
                options=all_ids,
                format_func=lambda x: f"{x} — {csf_df[csf_df['subcategory_id']==x]['subcategory_desc'].iloc[0][:70]}...",
                key="quick_assess_select",
            )

            if selected_sub:
                row = csf_df[csf_df["subcategory_id"] == selected_sub].iloc[0]
                current = get_assessment_for_sub(selected_sub)

                c1, c2 = st.columns([3, 2])
                with c1:
                    st.markdown(f"**{row['subcategory_id']}** — {row['subcategory_desc']}")
                    if row["implementation_examples"]:
                        with st.expander("Implementation Examples"):
                            st.markdown(row["implementation_examples"].replace("\n", "\n\n"))

                with c2:
                    new_status = st.select_slider(
                        "Implementation Status",
                        options=list(STATUS_LABELS.keys()),
                        value=current["status"],
                        format_func=lambda x: STATUS_LABELS[x],
                        key=f"status_{selected_sub}",
                    )
                    new_notes = st.text_area(
                        "Notes / Evidence / Gaps",
                        value=current["notes"],
                        height=120,
                        placeholder="e.g. Policy drafted, awaiting approval. Owner: CISO",
                        key=f"notes_{selected_sub}",
                    )
                    if st.button("💾 Save Assessment", type="primary", use_container_width=True, key=f"save_{selected_sub}"):
                        save_assessment(selected_sub, new_status, new_notes)
                        st.rerun()

    # ========== PAGE: SELF-ASSESSMENT ==========
    elif nav == "📝 Self-Assessment":
        st.subheader("Organization Self-Assessment")

        st.markdown(
            f"""
            **Organization:** {profile['org_name']}  
            Track your implementation maturity across all CSF 2.0 subcategories. 
            Status levels help prioritize roadmap and demonstrate progress to leadership / auditors.
            """
        )

        # Summary metrics
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Subcategories", stats["total"])
        m2.metric("Assessed", f"{stats['assessed']}", f"{stats['assessed']/stats['total']*100:.0f}%")
        m3.metric("Fully Implemented", stats["fully_implemented"])
        m4.metric("Avg Maturity", f"{stats['progress_pct']}%")

        st.divider()

        # Assessment table (read-only view with current status)
        assess_view = stats["merged"][["function_id", "category_id", "subcategory_id", "subcategory_desc", "status", "notes"]].copy()
        assess_view["Status"] = assess_view["status"].map(STATUS_LABELS)
        assess_view = assess_view.rename(columns={"subcategory_desc": "Description"})

        st.dataframe(
            assess_view[["function_id", "category_id", "subcategory_id", "Description", "Status", "notes"]],
            use_container_width=True,
            hide_index=True,
            column_config={
                "notes": st.column_config.TextColumn("Notes", width="medium"),
                "Status": st.column_config.TextColumn("Status"),
            },
        )

        st.caption("💡 Use the Browse & Search tab for easier hierarchical navigation and detailed editing.")

        # Bulk actions
        with st.expander("⚡ Bulk Actions"):
            if st.button("Mark all unassessed as 'Not Implemented' (status=1)"):
                for sid in csf_df["subcategory_id"]:
                    current = get_assessment_for_sub(sid)
                    if current["status"] == 0:
                        save_assessment(sid, 1, current["notes"])
                st.success("Bulk update complete!")
                st.rerun()

            if st.button("Reset ALL assessments (irreversible)", type="secondary"):
                conn = sqlite3.connect(DB_PATH)
                conn.execute("DELETE FROM assessments")
                conn.commit()
                conn.close()
                st.warning("All assessments have been reset.")
                st.rerun()

    # ========== PAGE: REPORTS & EXPORT ==========
    elif nav == "📈 Reports & Export":
        st.subheader("Implementation Reports & Export")

        st.markdown("### Progress by Function")
        func_progress = stats["per_function"].copy()
        func_progress["Function"] = func_progress["function_id"].map(lambda x: f"{x} — {FUNCTION_COLORS.get(x,{}).get('name',x)}")
        func_progress = func_progress[["Function", "subcats", "assessed", "progress_pct"]].rename(
            columns={"subcats": "Subcategories", "assessed": "Assessed", "progress_pct": "Progress %"}
        )

        st.dataframe(func_progress, use_container_width=True, hide_index=True)

        # Simple bar chart
        chart_data = stats["per_function"].set_index("function_id")["progress_pct"]
        st.bar_chart(chart_data, use_container_width=True, horizontal=True)

        st.divider()

        # Export
        st.markdown("### 📥 Export Your Assessment")

        export_df = stats["merged"][
            ["function_id", "category_id", "subcategory_id", "subcategory_desc", "status", "notes", "updated_at"]
        ].copy()
        export_df["Status Label"] = export_df["status"].map(STATUS_LABELS)
        export_df = export_df.drop(columns=["status"])

        csv = export_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="⬇️ Download Assessment as CSV",
            data=csv,
            file_name=f"CSF_2_0_Assessment_{profile['org_name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True,
        )

        st.caption("The CSV includes your current status, notes, and timestamps for audit / leadership reporting.")

        # Simple text summary
        with st.expander("Generate Executive Summary (Markdown)"):
            summary_md = f"""# CSF 2.0 Self-Assessment Summary

**Organization:** {profile['org_name']}  
**Date:** {datetime.now().strftime('%B %d, %Y')}  
**Overall Progress:** {stats['progress_pct']}% ({stats['assessed']}/{stats['total']} subcategories assessed)

## Key Highlights
- Fully Implemented: {stats['fully_implemented']}
- Average Maturity Score: {stats['progress_pct']}/100

## Progress by Function
"""
            for _, row in stats["per_function"].iterrows():
                summary_md += f"- **{row['function_id']}**: {row['progress_pct']}% ({row['assessed']}/{row['subcats']} assessed)\n"

            summary_md += "\n---\n*Generated with NIST CSF 2.0 Reference Tool*"
            st.code(summary_md, language="markdown")
            st.download_button("Download Summary .md", summary_md, file_name="CSF_Assessment_Summary.md")

    # ========== PAGE: ABOUT ==========
    else:
        st.subheader("About this Tool & NIST CSF 2.0")

        st.markdown(
            f"""
            This interactive application is built on the official **NIST Cybersecurity Framework (CSF) 2.0** Core, 
            downloaded from the CSF 2.0 Reference Tool on {intro.get('Generated Date', 'N/A')}.

            ### What is CSF 2.0?
            The NIST CSF provides a voluntary framework for organizations to manage cybersecurity risk. 
            Version 2.0 (2024) introduced the new **GOVERN (GV)** function and emphasizes supply chain, 
            governance, and measurable outcomes across six functions.

            ### Tool Features
            - Full hierarchical browsing of all 6 Functions, Categories, and 185 Subcategories
            - Implementation examples for each subcategory
            - Persistent self-assessment with 5-level maturity scoring
            - Progress dashboards and exportable reports (CSV + Markdown)
            - Organization profile for contextual notes

            ### Data Notes
            - 79 subcategories are marked **Withdrawn** (mostly from IDENTIFY) as they were consolidated into GOVERN or other areas in v2.0.
            - Informative References column is empty in this export (mappings to SP 800-53, ISO 27001 etc. are available on the official NIST site).
            """
        )

        with st.expander("Official Sources & Feedback"):
            st.markdown(
                """
                - Official NIST CSF 2.0: https://www.nist.gov/cyberframework
                - CSF 2.0 Reference Tool: https://csrc.nist.gov/projects/cybersecurity-framework/csf-2-0-reference-tool
                - Feedback to NIST: cprt@nist.gov (as noted in the source file)

                This tool is for internal use / education and is not an official NIST product.
                """
            )

        st.success("Built as a full-stack reference & assessment application • Streamlit + SQLite backend")

    # Footer
    st.divider()
    st.caption(
        "NIST CSF 2.0 Explorer • Data source: cprt_CSF_2_0_0_06-17-2026.xlsx • "
        "Not affiliated with NIST • For reference and internal assessment use only"
    )


if __name__ == "__main__":
    main()
